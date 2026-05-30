import asyncio
import io
import json
import logging
import os
import traceback
import uuid
from contextlib import asynccontextmanager
from typing import List, Optional

from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles

# Must load .env before importing agent (which reads env vars at init time)
load_dotenv()

from agent import WebTestingAgent
from database import Database

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Registry of active test runs: run_id → asyncio.Event (set to signal stop)
_stop_events: dict[str, asyncio.Event] = {}


async def extract_file_text(file: UploadFile) -> str:
    content = await file.read()
    name = (file.filename or "").lower()
    try:
        if name.endswith((".txt", ".csv", ".md")):
            return content.decode("utf-8", errors="replace")
        if name.endswith(".pdf"):
            import pdfplumber
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                return "\n".join(page.extract_text() or "" for page in pdf.pages)
        if name.endswith(".docx"):
            from docx import Document
            doc = Document(io.BytesIO(content))
            return "\n".join(p.text for p in doc.paragraphs)
        if name.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            rows = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    line = "\t".join("" if c is None else str(c) for c in row)
                    if line.strip():
                        rows.append(line)
            return "\n".join(rows)
    except Exception as exc:
        logger.warning("Could not parse %s: %s", file.filename, exc)
    return ""

# Use absolute paths so db and screenshots always land in backend/ folder
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "web_tester.db")
SCREENSHOTS_DIR = os.path.join(BASE_DIR, "screenshots")

db = Database(db_path=DB_PATH)

os.makedirs(SCREENSHOTS_DIR, exist_ok=True)


def parse_keys_list(api_keys: Optional[str]) -> Optional[List[str]]:
    if not api_keys:
        return None
    try:
        keys_list = json.loads(api_keys)
        if isinstance(keys_list, list):
            return [str(k).strip() for k in keys_list if str(k).strip()]
        return [str(keys_list).strip()]
    except json.JSONDecodeError:
        normalized = api_keys.replace("\r\n", ",").replace("\n", ",")
        return [k.strip() for k in normalized.split(",") if k.strip()]


@asynccontextmanager
async def lifespan(app: FastAPI):
    db.init()
    yield


app = FastAPI(title="AI Web Tester", version="1.0.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/screenshots", StaticFiles(directory=SCREENSHOTS_DIR), name="screenshots")


@app.get("/health")
async def health():
    configured = {
        "anthropic": bool(os.getenv("ANTHROPIC_API_KEY")),
        "openai": bool(os.getenv("OPENAI_API_KEY")),
        "gemini": bool(os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")),
        "groq": bool(os.getenv("GROQ_API_KEY")),
    }
    return {"status": "ok", "service": "AI Web Tester", "api_keys_configured": configured}


@app.post("/stop/{run_id}")
async def stop_test(run_id: str):
    event = _stop_events.get(run_id)
    if event:
        event.set()
        return {"message": "Stop signal sent"}
    # Test may have already finished — not an error
    return {"message": "Test already completed or not found"}


@app.post("/test")
async def run_test(
    url: str = Form(...),
    goal: str = Form(...),
    credentials: Optional[str] = Form(None),
    files: List[UploadFile] = File(default=[]),
    run_id: Optional[str] = Form(None),
    provider: Optional[str] = Form(None),
    model: Optional[str] = Form(None),
    api_keys: Optional[str] = Form(None),
    custom_base_url: Optional[str] = Form(None),
    all_provider_keys: Optional[str] = Form(None),
):
    if not url.strip():
        raise HTTPException(status_code=400, detail="URL is required")
    if not goal.strip():
        raise HTTPException(status_code=400, detail="Goal is required")

    keys_list = parse_keys_list(api_keys)
    
    all_keys_dict: Optional[dict] = None
    if all_provider_keys:
        try:
            all_keys_dict = json.loads(all_provider_keys)
        except json.JSONDecodeError:
            pass

    credentials_dict: Optional[dict] = None
    if credentials:
        try:
            credentials_dict = json.loads(credentials)
        except json.JSONDecodeError:
            pass

    effective_goal = goal.strip()
    real_files = [f for f in (files or []) if f.filename]
    if real_files:
        file_sections = []
        for f in real_files:
            text = await extract_file_text(f)
            if text.strip():
                file_sections.append(f"=== {f.filename} ===\n{text.strip()}")
        if file_sections:
            effective_goal += "\n\nInstructions from attached files:\n" + "\n\n".join(file_sections)

    active_run_id = run_id or str(uuid.uuid4())
    stop_event = asyncio.Event()
    _stop_events[active_run_id] = stop_event
    try:
        agent = WebTestingAgent(
            provider=provider,
            model=model,
            api_keys=keys_list,
            custom_base_url=custom_base_url,
            provider_keys=all_keys_dict,
        )
        result = await agent.run(
            url=url.strip(),
            goal=effective_goal,
            credentials=credentials_dict,
            stop_event=stop_event,
        )
    except Exception as exc:
        logger.error("Agent error: %s", traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Agent error: {exc}")
    finally:
        _stop_events.pop(active_run_id, None)

    test_id = await asyncio.to_thread(db.save_test_run, url, goal, result)
    result["test_id"] = test_id

    for step in result.get("steps", []):
        if step.get("screenshot"):
            fname = os.path.basename(step["screenshot"])
            step["screenshot_url"] = f"/screenshots/{fname}"

    logger.info("Test complete: %s | %s", result.get("final_verdict"), url)
    return result


@app.post("/test-stream")
async def run_test_stream(
    url: str = Form(...),
    goal: str = Form(...),
    credentials: Optional[str] = Form(None),
    files: List[UploadFile] = File(default=[]),
    headed: str = Form("false"),
    slow_mo: int = Form(0),
    run_id: Optional[str] = Form(None),
    provider: Optional[str] = Form(None),
    model: Optional[str] = Form(None),
    api_keys: Optional[str] = Form(None),
    custom_base_url: Optional[str] = Form(None),
    all_provider_keys: Optional[str] = Form(None),
):
    if not url.strip():
        raise HTTPException(status_code=400, detail="URL is required")
    if not goal.strip():
        raise HTTPException(status_code=400, detail="Goal is required")

    keys_list = parse_keys_list(api_keys)
    
    all_keys_dict: Optional[dict] = None
    if all_provider_keys:
        try:
            all_keys_dict = json.loads(all_provider_keys)
        except json.JSONDecodeError:
            pass

    is_headed = headed.lower() in ("true", "1", "yes")

    credentials_dict: Optional[dict] = None
    if credentials:
        try:
            credentials_dict = json.loads(credentials)
        except json.JSONDecodeError:
            pass

    # Read file contents eagerly before the streaming generator starts
    effective_goal = goal.strip()
    real_files = [f for f in (files or []) if f.filename]
    if real_files:
        file_sections = []
        for f in real_files:
            text = await extract_file_text(f)
            if text.strip():
                file_sections.append(f"=== {f.filename} ===\n{text.strip()}")
        if file_sections:
            effective_goal += "\n\nInstructions from attached files:\n" + "\n\n".join(file_sections)

    active_run_id = run_id or str(uuid.uuid4())
    stop_event = asyncio.Event()
    _stop_events[active_run_id] = stop_event

    async def event_generator():
        # Send run_id immediately so the frontend can wire up the stop button
        yield f"data: {json.dumps({'type': 'init', 'run_id': active_run_id})}\n\n"

        queue: asyncio.Queue = asyncio.Queue()

        async def on_log(event: dict):
            await queue.put(event)

        async def run_agent_task():
            try:
                agent = WebTestingAgent(
                    provider=provider,
                    model=model,
                    api_keys=keys_list,
                    custom_base_url=custom_base_url,
                    provider_keys=all_keys_dict,
                )
                result = await agent.run(
                    url=url.strip(),
                    goal=effective_goal,
                    credentials=credentials_dict,
                    headless=not is_headed,
                    slow_mo=slow_mo,
                    on_log=on_log,
                    stop_event=stop_event,
                )
                test_id = await asyncio.to_thread(db.save_test_run, url, goal, result)
                result["test_id"] = test_id
                for step in result.get("steps", []):
                    if step.get("screenshot"):
                        fname = os.path.basename(step["screenshot"])
                        step["screenshot_url"] = f"/screenshots/{fname}"
                await queue.put({"type": "result", "data": result})
            except Exception as exc:
                logger.error("Stream agent error: %s", traceback.format_exc())
                await queue.put({"type": "error", "message": str(exc)})
            finally:
                _stop_events.pop(active_run_id, None)
                await queue.put(None)

        task = asyncio.create_task(run_agent_task())
        try:
            while True:
                event = await asyncio.wait_for(queue.get(), timeout=360)
                if event is None:
                    break
                yield f"data: {json.dumps(event)}\n\n"
        except asyncio.TimeoutError:
            yield f"data: {json.dumps({'type': 'error', 'message': 'Test timed out after 6 minutes'})}\n\n"
        finally:
            if not task.done():
                task.cancel()

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/tests")
async def get_tests(limit: int = 20, offset: int = 0):
    return await asyncio.to_thread(db.get_test_runs, limit, offset)


@app.get("/tests/{test_id}")
async def get_test(test_id: str):
    result = await asyncio.to_thread(db.get_test_run, test_id)
    if not result:
        raise HTTPException(status_code=404, detail="Test not found")
    return result


@app.delete("/tests/{test_id}")
async def delete_test(test_id: str):
    deleted = await asyncio.to_thread(db.delete_test_run, test_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Test not found")
    return {"message": "Deleted successfully"}


if __name__ == "__main__":
    import uvicorn
    # reload=False avoids subprocess issues with .env and Playwright on Windows
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)
